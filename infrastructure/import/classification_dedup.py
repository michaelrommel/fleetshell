#!/usr/bin/env python3
"""
classification_dedup.py -- turn the legacy Data Classification spreadsheet into a
deduped, name-keyed artifact (classification.json) that the importer re-applies
by product/family NAME (never UUID), so it survives an old_database reload.

DB-free and deterministic. See docs/data_classification.md.

Model recap:
  * A rule   = (filename regex, set of data-class codes).
  * Support  = the set of targets (products, or families for CT/MR) carrying it.
  * Rule Set = all rules with IDENTICAL support -> assigned to exactly that
               target set. This is the minimal, lossless "best shape for reuse".
  * XP/AX/NM/PACS/CP: sheet "Product" column = a real product  -> product-level.
  * CT/MR            : sheet "Product" column = a platform/family -> family-level
                       (dormant until product.family is populated to match).

Usage:
  python classification_dedup.py [xlsx] \
      [--out classification.json] [--families product_families.json]
"""
import argparse, json, os, re, sys
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLSX = os.path.join(HERE, "old_database", "Data_Classification_fleetshell.xlsx")
DEFAULT_CIM = os.path.join(HERE, "old_database", "Data_Classification_CIM_fleetshell.xlsx")

# The CIM/helpdesk file: one flat, named set of platform-generic rules, assigned
# modality-wide into each modality present in the CIM sheet (Product='all_products').
CIM_SET_NAME = "Helpdesk/CIM Files"

# Modalities whose "Product" column is actually a platform/family.
FAMILY_MODALITIES = {"CT", "MR"}

# Sheet class-column header (stripped) -> our data_class code, in canonical order.
CLASS_HEADERS = [
    ("Protected Health Information (PHI)", "PHI"),
    ("Utilization & Performance Data",     "UPD"),
    ("Result Data",                        "RD"),
    ("Personal Identifiable Information",  "PII"),
    ("Asset & Configuration Data",         "ACD"),
    ("Device Service History",             "DSH"),
    ("Technical Status Data",              "TSD"),
    ("Smart Technical Data",               "STD"),
]
CLASS_MAP = {h: c for h, c in CLASS_HEADERS}

HDR_MODALITY = "Modality"
HDR_PRODUCT  = "Product"
HDR_REGEX    = "Filename Regular Expression (text input)"

# A cell counts as "checked" unless empty or an explicit negative.
NEG = {"", "no", "n", "-", "none", "na", "n/a", "0", "false", "x-no"}

# Generic tokens we avoid picking as a set NAME when a better one exists.
NAME_STOP = {"zip", "tmp", "txt", "log", "logs", "xml", "gz", "csv", "tar",
             "file", "files", "data", "info", "the", "and"}


def is_checked(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s not in NEG


def tokenize(regex: str):
    """Alphanumeric tokens (>=3 chars, containing a letter) from a regex."""
    out = []
    for t in re.split(r"[^A-Za-z0-9]+", regex):
        if len(t) >= 3 and any(ch.isalpha() for ch in t):
            out.append(t)
    return out


def name_for(regexes, used: set) -> str:
    """Heuristic, deterministic set name from the common regex token."""
    token_sets = [set(tokenize(r)) for r in regexes]
    common = set.intersection(*token_sets) if token_sets and all(token_sets) else set()

    def rank(tokens):
        # prefer non-stop, then longer, then alphabetical
        return sorted(tokens, key=lambda t: (t.lower() in NAME_STOP, -len(t), t.lower()))

    base = None
    if common:
        base = rank(common)[0]
    else:
        c = Counter(t for ts in token_sets for t in ts)
        if c:
            # most frequent; tie-break by rank
            top = max(c.values())
            base = rank([t for t, n in c.items() if n == top])[0]
    if not base:
        base = "shared"

    name, i = base, 1
    while name in used:
        i += 1
        name = f"{base}-{i}"
    used.add(name)
    return name


def load_sheet(ws):
    """Yield (modality, target_name, regex, codes_tuple) for a modality sheet.
    The modality is taken from the row's Modality COLUMN (not the sheet title),
    so a row authored on one BU's sheet but owned by another tree (e.g. the AX
    sheet's Cios rows whose Modality column = XP) routes to the right modality.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [(str(c).strip() if c is not None else "") for c in rows[0]]

    def col(name):
        return header.index(name) if name in header else None

    ci_mod = col(HDR_MODALITY)
    ci_prod = col(HDR_PRODUCT)
    ci_regex = col(HDR_REGEX)
    if ci_prod is None or ci_regex is None:
        return  # not a modality data sheet
    class_cols = [(header.index(h), code) for h, code in CLASS_HEADERS if h in header]
    sheet_modality = ws.title.strip()

    for r in rows[1:]:
        def cell(i):
            return r[i] if i is not None and i < len(r) else None
        modality = cell(ci_mod)
        modality = str(modality).strip() if modality is not None else ""
        modality = modality or sheet_modality      # blank column -> sheet name
        target = cell(ci_prod)
        regex = cell(ci_regex)
        target = str(target).strip() if target is not None else ""
        regex = str(regex).strip() if regex is not None else ""
        if not target or not regex:
            continue
        codes = tuple(code for ci, code in class_cols if is_checked(cell(ci)))
        if not codes:
            continue  # a rule with no class is meaningless -> skip junk/sub-header rows
        yield modality, target, regex, codes


def load_cim(path):
    """Read the CIM/helpdesk sheet -> {modality: [ {regex, codes}, ... ]}.
    Product is always 'all_products' (modality-wide); rules are a flat list per
    modality (deduped, order preserved). Columns located by header name."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [(str(c).strip() if c is not None else "") for c in rows[0]]
    if HDR_MODALITY not in header or HDR_REGEX not in header:
        return {}
    ci_mod = header.index(HDR_MODALITY)
    ci_regex = header.index(HDR_REGEX)
    class_cols = [(header.index(h), code) for h, code in CLASS_HEADERS if h in header]

    by_mod = {}
    for r in rows[1:]:
        def cell(i):
            return r[i] if i is not None and i < len(r) else None
        modality = str(cell(ci_mod)).strip() if cell(ci_mod) is not None else ""
        regex = str(cell(ci_regex)).strip() if cell(ci_regex) is not None else ""
        if not modality or not regex:
            continue
        codes = tuple(code for ci, code in class_cols if is_checked(cell(ci)))
        if not codes:
            continue
        seen, lst = by_mod.setdefault(modality, (set(), []))
        key = (regex, codes)
        if key not in seen:
            seen.add(key)
            lst.append({"regex": regex, "codes": list(codes)})
    return {m: lst for m, (seen, lst) in by_mod.items()}


def factor_modality(pairs):
    """
    pairs: iterable of (target, regex, codes_tuple).
    Returns (sets, stats). sets = list of dicts {rules, targets}.
    Rules with identical target-support are grouped into one set.
    """
    # rule -> set of targets (dedup identical (regex,codes) within a target too)
    support = defaultdict(set)              # (regex, codes) -> {targets}
    for target, regex, codes in pairs:
        support[(regex, codes)].add(target)

    # group rules by identical support signature
    by_sig = defaultdict(list)              # frozenset(targets) -> [(regex,codes)]
    for rule, targets in support.items():
        by_sig[frozenset(targets)].append(rule)

    sets = []
    for targets, rules in by_sig.items():
        # stable ordering of rules
        rules_sorted = sorted(rules, key=lambda rc: rc[0])
        sets.append({"targets": sorted(targets), "rules": rules_sorted})
    # largest support first (shared bundles on top), then by rule count
    sets.sort(key=lambda s: (-len(s["targets"]), -len(s["rules"]), s["rules"][0][0]))
    return sets


def main():
    import openpyxl
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX)
    ap.add_argument("--cim", default=DEFAULT_CIM, help="CIM/helpdesk sheet (Product='all_products')")
    ap.add_argument("--out", default=os.path.join(HERE, "classification.json"))
    ap.add_argument("--families", default=os.path.join(HERE, "product_families.json"))
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, read_only=True, data_only=True)
    artifact = {"version": 1, "modalities": {}}
    families_scaffold = {}
    print(f"reading {a.xlsx}")

    # Phase 1: collect (target, regex, codes) buckets keyed by the row's Modality
    # COLUMN across every modality data sheet (so cross-tree rows route correctly).
    pairs_by_mod = defaultdict(list)
    for ws in wb.worksheets:
        rows0 = next(ws.iter_rows(values_only=True), None)
        if not rows0:
            continue
        head = [(str(c).strip() if c is not None else "") for c in rows0]
        if HDR_MODALITY not in head or HDR_PRODUCT not in head or HDR_REGEX not in head:
            continue  # meta sheet
        for modality, target, regex, codes in load_sheet(ws):
            pairs_by_mod[modality].append((target, regex, codes))

    # Phase 2: factor each modality bucket independently.
    for modality in sorted(pairs_by_mod):
        pairs = pairs_by_mod[modality]
        is_family = modality in FAMILY_MODALITIES
        sets = factor_modality(pairs)

        used_names = set()
        out_sets = []
        for s in sets:
            nm = name_for([rgx for rgx, _ in s["rules"]], used_names)
            rules = [{"regex": rgx, "codes": list(codes)} for rgx, codes in s["rules"]]
            assign = {"modality_wide": False, "products": [], "families": []}
            if is_family:
                assign["families"] = s["targets"]
            else:
                assign["products"] = s["targets"]
            out_sets.append({"name": nm, "description": "", "rules": rules, "assign": assign})

        artifact["modalities"][modality] = {"family_based": is_family, "sets": out_sets}
        n_rules = sum(len(s["rules"]) for s in out_sets)
        n_tgt = len({t for s in sets for t in s["targets"]})
        print(f"  {modality:6} : {len(out_sets):3} sets, {n_rules:4} rules, "
              f"{n_tgt:3} {'families' if is_family else 'products'}")
        if is_family:
            families_scaffold[modality] = sorted({t for s in sets for t in s["targets"]})

    # CIM/helpdesk: one flat named set assigned modality-wide, spread into each
    # modality present in the CIM sheet (Product='all_products').
    if a.cim and os.path.exists(a.cim):
        cim = load_cim(a.cim)
        for modality, rules in cim.items():
            block = artifact["modalities"].setdefault(modality, {"family_based": False, "sets": []})
            block["sets"] = [s for s in block["sets"] if s["name"] != CIM_SET_NAME]  # idempotent
            block["sets"].append({
                "name": CIM_SET_NAME, "description": "", "rules": rules,
                "assign": {"modality_wide": True, "products": [], "families": []},
            })
            print(f"  {modality:6} : + '{CIM_SET_NAME}' ({len(rules)} rules, modality-wide)")
    elif a.cim:
        print(f"  CIM sheet {a.cim} absent -- skipped")

    with open(a.out, "w") as f:
        json.dump(artifact, f, indent=1, ensure_ascii=False)
        f.write("\n")
    print(f"wrote {a.out}")

    if families_scaffold:
        # product_families.json is the user-authored mapping (family -> product
        # names) read by load.py --stage families. NEVER clobber it: write a
        # template only if it does not exist; otherwise just report family
        # strings from the sheet that are not yet mapped there.
        fam_path = a.families
        if os.path.exists(fam_path):
            try:
                existing = json.load(open(fam_path))
            except Exception:
                existing = {}
            missing = []
            for m, fams in families_scaffold.items():
                have = set((existing.get(m) or {}).keys())
                missing += [f"{m} / {fam}" for fam in fams if fam not in have]
            if missing:
                print(f"  {fam_path} exists; UNMAPPED family strings: {', '.join(missing)}")
            else:
                print(f"  {fam_path} exists; all sheet family strings are mapped")
        else:
            template = {m: {fam: [] for fam in fams} for m, fams in families_scaffold.items()}
            with open(fam_path, "w") as f:
                json.dump(template, f, indent=1, ensure_ascii=False)
                f.write("\n")
            print(f"wrote template {fam_path} (fill in product names per family)")


if __name__ == "__main__":
    main()
