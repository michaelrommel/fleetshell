#!/usr/bin/env python3
"""dtm_dedup.py -- Data Transfer Matrix workbooks -> committed dtm.json.

Parses the per-country SRS Data Transfer Matrix spreadsheets in
old_database/DataTransferMatrix/ into a single, ISO-keyed, reload-survivable
artifact (dtm.json) that load.py --stage dtm re-applies. Like classification.json
it is name/code-keyed (never DB UUIDs), so it survives a full reload and a
DB->file round-trip can later capture Country-Manager edits.

Model (see docs/data_transfer_matrix.md):
  * One workbook = one FROM-country x variant (Standard/Strict).
  * Grid rows = destination country, columns = data class, cell in
    {permitted, denied} (unknown/contract are defined but unused in practice).
  * We keep ONLY 'denied' cells (denial-list, default = permit).

Output shape:
  {
    "variants": {"STD": "Standard", "STR": "Strict"},
    "matrices": {
      "DE": { "STD": {"default": "permit",
                      "deny": {"CN": ["DSH","RS","RDS","UPD","PII","PHI"], ...}},
              "STR": {...} },
      ...
    }
  }

Usage:
    python dtm_dedup.py            # writes dtm.json + prints a resolution report
"""
from __future__ import annotations
import json, os, sys, glob
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "old_database", "DataTransferMatrix")
OUT = os.path.join(HERE, "dtm.json")

# FROM-country filename token -> ISO-3166 alpha-2 (the deployed set).
FILE_COUNTRY_ISO = {
    "Finland": "FI", "Germany": "DE", "Italy": "IT",
    "Japan": "JP", "Norway": "NO", "USA": "US",
}
VARIANT_CODE = {"Standard": "STD", "Strict": "STR"}
VARIANT_LABEL = {"STD": "Standard", "STR": "Strict"}

# Matrix column display name -> canonical data_class code (must match
# migrate_dtm.sql). Both "Personal Data" (older) and "Personal Identifiable
# Information" (newer) are the same class (MRS_DC_IO_REC_PersonalData) -> PII.
NAME_TO_CODE = {
    "Technical Status Data": "TSD",
    "Asset And Configuration Data": "ACD",
    "Result Data": "RD",
    "Smart Technical Data": "STD",
    "Device Service History": "DSH",
    "Reactive Sessions": "RS",
    "Remote Desktop Sharing": "RDS",
    "RSWD Status Feedbacks": "RSWD",
    "Utilization Performance Data": "UPD",
    "Personal Data": "PII",
    "Personal Identifiable Information": "PII",
    "Protected Health Information": "PHI",
    "Software Distribution": "SWD",
    "Software Distribution Orders": "SWDO",
    "Virus Pattern Distribution": "VPD",
    "Remote Option Distribution": "ROD",
    "Smart Data": "SMD",
}

# Retired data classes: some older sheets still carry these columns, but the
# classes were removed over time. Ignore them entirely (do NOT record denials).
IGNORED_CLASS_NAMES = {
    "Result Data 3rd Party",
    "Utilization Performance Data 3rd Party",
    "Pseudonymized Health Information",
    "Generic File Transfer",
}
STATUS = {"permitted", "denied", "unknown", "contract"}


def parse_workbook(path: str, report: Counter) -> tuple[str, str, dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # FROM country + variant from the filename.
    base = os.path.basename(path).replace("SRS-DataTransferMatrix-", "").replace(".xlsx", "")
    country_tok, variant_tok = base.rsplit("-", 1)
    from_iso = FILE_COUNTRY_ISO.get(country_tok)
    variant = VARIANT_CODE.get(variant_tok)
    if not from_iso or not variant:
        raise SystemExit(f"unknown FROM/variant in filename: {base!r}")

    # Country long-name -> ISO from the Parameters "Country List".
    name_to_iso: dict[str, str] = {}
    for r in wb["Parameters"].iter_rows(values_only=True):
        for j, c in enumerate(r):
            if isinstance(c, str) and j + 1 < len(r):
                nxt = r[j + 1]
                if isinstance(nxt, str) and len(nxt) == 2 and nxt.isalpha() and nxt.isupper():
                    name_to_iso.setdefault(c.strip(), nxt)

    ws = wb["DTM"]
    rows = list(ws.iter_rows(values_only=True))
    header_cols: list[str] | None = None
    deny: dict[str, list[str]] = {}
    for r in rows:
        c0 = r[0]
        if c0 == "Data Class":
            header_cols = [c for c in r[4:21] if c]
            continue
        if header_cols is None:
            continue
        # a destination-country row: col0 is a known country and cells are statuses
        if isinstance(c0, str) and any(v in STATUS for v in r[4:21]):
            to_iso = name_to_iso.get(c0.strip())
            if not to_iso:
                report[f"unresolved_dest:{c0}"] += 1
                continue
            denied_codes = []
            for name, val in zip(header_cols, r[4 : 4 + len(header_cols)]):
                if val == "denied":
                    if name in IGNORED_CLASS_NAMES:
                        continue                      # retired class column -> ignore
                    code = NAME_TO_CODE.get(name)
                    if not code:
                        report[f"unknown_class:{name}"] += 1
                        continue
                    denied_codes.append(code)
                elif val not in ("permitted", None):
                    report[f"nonbinary_cell:{val}"] += 1
            if denied_codes:
                # dedup + stable order
                deny[to_iso] = sorted(set(denied_codes))
    return from_iso, variant, {"default": "permit", "deny": deny}


def main() -> None:
    report: Counter = Counter()
    matrices: dict[str, dict] = {}
    files = sorted(glob.glob(os.path.join(SRC, "*.xlsx")))
    if not files:
        raise SystemExit(f"no workbooks in {SRC}")
    for path in files:
        from_iso, variant, block = parse_workbook(path, report)
        matrices.setdefault(from_iso, {})[variant] = block
        ndeny = sum(len(v) for v in block["deny"].values())
        print(f"  {os.path.basename(path):42} {from_iso}/{variant}: "
              f"{len(block['deny'])} dest with denials, {ndeny} denied cells")

    out = {"variants": VARIANT_LABEL, "matrices": matrices}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False, sort_keys=True)
    print(f"\nwrote {os.path.relpath(OUT, HERE)}: {len(matrices)} FROM-countries")

    if report:
        print("\n--- resolution report (review) ---")
        for k, n in sorted(report.items()):
            print(f"  {k}: {n}")


if __name__ == "__main__":
    main()
