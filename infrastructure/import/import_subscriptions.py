#!/usr/bin/env python3
"""import_subscriptions.py -- load the legacy File Subscriptions overview into the
new master-data tables (subscriber_server / subscription / subscription_server).

Source: old_database/subscription_overview_2026-05-05.xlsx (gitignored, like the
rest of old_database). One denormalized row per (server x subscription); the same
subscription line repeats across every server it is attached to.

Columns:
  ipaddress subscriber country subscription modality productmodel namepattern
  negated quality_file_subscription modificationdate annotation creatinguser
  modifyinguser

Mapping into the new model (see migrate_file_subscriptions.sql):
  subscriber_server : name=subscriber, ip_address=ipaddress, country=<ISO of `country`>,
                      use_case=DEFAULT_USE_CASE, activated=DEFAULT_ACTIVATED,
                      delivery_method=DEFAULT_DELIVERY.  (No delivery/auth/root
                      existed in legacy -- admins fill those in later.)
  subscription      : name=subscription, modality_id=<by name>,
                      product_id=<by name, the part after ' / '>, pattern=namepattern,
                      negate=(negated=='negated').
  subscriber_server.comment : the anonymized `annotation` (contacts/purpose). It
                      is consistent per server in the source, so it maps cleanly
                      onto the server rather than each subscription.
  subscription_server: the (subscriber, subscription) attach matrix.

Anonymization honors the SAME ANONYMIZE switch as load.py (default on). Only the
free-text `annotation` carries PII (names/emails); it is replaced with the shared
ANON_TEXT_PLACEHOLDER when ANONYMIZE=1, and passes through raw when ANONYMIZE=0
(production take-over with real data). The legacy IP is imported as-is either way.

Idempotent: upserts by the unique `name` on both entities, so a re-run (or a
post-reload run, since a `product` TRUNCATE CASCADE empties subscription +
subscription_server) rebuilds cleanly and re-resolves the product/modality FKs.

Env:
  IMPORT_GLOBAL_DSN   e.g. "host=localhost port=5432 dbname=fleetshell user=fsadmin password=... sslmode=require"
  ANONYMIZE           1 (default) = fake PII; 0 = raw (production take-over)

Usage:
  IMPORT_GLOBAL_DSN=... python import_subscriptions.py
  IMPORT_GLOBAL_DSN=... python import_subscriptions.py --file path/to.xlsx --dry-run
"""

from __future__ import annotations
import argparse, os, sys
from collections import defaultdict

import psycopg
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.join(HERE, "old_database", "subscription_overview_2026-05-05.xlsx")

# Same switch + placeholder semantics as load.py.
ANONYMIZE = os.environ.get("ANONYMIZE", "1").strip().lower() not in ("0", "false", "no", "off")
ANON_TEXT_PLACEHOLDER = "<content was anonymized during seeding>"

# Import defaults for fields absent from the legacy export (confirmed with owner).
DEFAULT_DELIVERY = "scp"          # legacy IP-based FTP landing zones
DEFAULT_USE_CASE = "compliance"
DEFAULT_ACTIVATED = True


def anon_annotation(raw: str | None) -> str | None:
    """Free-text contact/purpose note: fixed placeholder when anonymizing (it is
    PII and cannot be structurally faked), raw text otherwise. Empty -> NULL."""
    v = (raw or "").strip()
    if not v:
        return None
    return ANON_TEXT_PLACEHOLDER if ANONYMIZE else v


def load_lookups(g: psycopg.Connection):
    """Name-keyed resolvers from the live product tree + region country ISO."""
    with g.cursor() as c:
        c.execute("SELECT name, id FROM product WHERE kind='modality' AND name<>''")
        modality_by_name = {n: i for n, i in c.fetchall()}
        # Product names are globally unique across modalities (verified), so a
        # bare name-keyed map is unambiguous.
        c.execute("SELECT name, id FROM product WHERE kind='product' AND name<>''")
        product_by_name: dict[str, str] = {}
        for n, i in c.fetchall():
            product_by_name.setdefault(n, i)
        c.execute("SELECT iso, name FROM region WHERE nlevel(path)=2 AND iso IS NOT NULL")
        country_rows = c.fetchall()
    # country name -> ISO: exact match first, then a prefix match ("United States"
    # -> "United States of America"). Deterministic (sorted).
    iso_by_exact = {name: iso for iso, name in country_rows}
    def country_iso(name: str | None) -> str | None:
        v = (name or "").strip()
        if not v:
            return None
        if v in iso_by_exact:
            return iso_by_exact[v]
        for iso, rn in sorted(country_rows, key=lambda t: t[1]):
            if rn.startswith(v):
                return iso
        return None
    return modality_by_name, product_by_name, country_iso


def read_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    idx = {name: i for i, name in enumerate(hdr)}
    need = ["ipaddress", "subscriber", "country", "subscription", "modality",
            "productmodel", "namepattern", "negated", "annotation", "modificationdate"]
    for col in need:
        if col not in idx:
            sys.exit(f"missing expected column '{col}' in {path}")
    out = []
    for r in it:
        if not r or all(v is None for v in r):
            continue
        out.append({col: r[idx[col]] for col in need})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=DEFAULT_FILE)
    ap.add_argument("--dry-run", action="store_true", help="resolve + report, do not write")
    a = ap.parse_args()

    if not os.path.exists(a.file):
        sys.exit(f"source not found: {a.file}")
    dsn = os.environ.get("IMPORT_GLOBAL_DSN")
    if not dsn:
        sys.exit("IMPORT_GLOBAL_DSN is not set")

    rows = read_rows(a.file)
    print(f"read {len(rows)} rows from {os.path.basename(a.file)}  (ANONYMIZE={'on' if ANONYMIZE else 'off'})")

    g = psycopg.connect(dsn)
    modality_by_name, product_by_name, country_iso = load_lookups(g)

    # --- collapse the denormalized rows -------------------------------------
    servers: dict[str, dict] = {}          # subscriber name -> fields
    subs: dict[str, dict] = {}             # subscription name -> fields
    attach: set[tuple[str, str]] = set()   # (subscriber, subscription)
    missing_mod, missing_prod, missing_country = set(), set(), set()

    for r in rows:
        sname = (r["subscriber"] or "").strip()
        subname = (r["subscription"] or "").strip()
        if not sname:
            continue

        # server (first IP/country wins; each name maps to one IP in the source.
        # annotation is consistent per server, so keep the first non-empty one).
        srv = servers.get(sname)
        if srv is None:
            iso = country_iso(r["country"])
            if r["country"] and iso is None:
                missing_country.add(str(r["country"]))
            servers[sname] = srv = {
                "ip_address": (r["ipaddress"] or "").strip() or None,
                "country": iso,
                "annotation": (r["annotation"] or "").strip(),
            }
        elif not srv["annotation"]:
            srv["annotation"] = (r["annotation"] or "").strip()

        # a server row without a subscription (e.g. an empty landing zone) still
        # registers the server above; nothing more to do for this row.
        if not subname:
            continue

        # subscription (rows are internally consistent per subscription name)
        pm = (r["productmodel"] or "").strip()
        product_name = pm.split(" / ", 1)[1] if " / " in pm else None
        mod_id = modality_by_name.get((r["modality"] or "").strip()) if r["modality"] else None
        prod_id = product_by_name.get(product_name) if product_name else None
        if r["modality"] and mod_id is None:
            missing_mod.add(str(r["modality"]))
        if product_name and prod_id is None:
            missing_prod.add(product_name)

        if subname not in subs:
            subs[subname] = {
                "modality_id": mod_id,
                "product_id": prod_id,
                "pattern": (r["namepattern"] or "").strip(),
                "negate": (r["negated"] or "").strip().lower() == "negated",
            }

        attach.add((sname, subname))

    print(f"collapsed -> {len(servers)} servers, {len(subs)} subscriptions, {len(attach)} attachments")
    if missing_country:
        print("  WARN unmapped countries:", ", ".join(sorted(missing_country)))
    if missing_mod:
        print("  WARN unmapped modalities:", ", ".join(sorted(missing_mod)))
    if missing_prod:
        print(f"  WARN {len(missing_prod)} unmapped product name(s):", ", ".join(sorted(missing_prod)))

    if a.dry_run:
        print("dry-run: no writes")
        g.close()
        return

    # --- write (upsert by unique name) --------------------------------------
    with g.cursor() as c:
        server_id: dict[str, str] = {}
        for name, f in servers.items():
            c.execute(
                """INSERT INTO subscriber_server
                     (name, ip_address, country, use_case, comment, activated, delivery_method)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT (name) DO UPDATE SET
                     ip_address=EXCLUDED.ip_address, country=EXCLUDED.country,
                     use_case=EXCLUDED.use_case, comment=EXCLUDED.comment,
                     activated=EXCLUDED.activated,
                     delivery_method=EXCLUDED.delivery_method, updated_at=now()
                   RETURNING id""",
                (name, f["ip_address"], f["country"], DEFAULT_USE_CASE,
                 anon_annotation(f["annotation"]), DEFAULT_ACTIVATED, DEFAULT_DELIVERY),
            )
            server_id[name] = c.fetchone()[0]

        sub_id: dict[str, str] = {}
        for name, f in subs.items():
            c.execute(
                """INSERT INTO subscription
                     (name, modality_id, product_id, pattern, negate)
                   VALUES (%s,%s,%s,%s,%s)
                   ON CONFLICT (name) DO UPDATE SET
                     modality_id=EXCLUDED.modality_id, product_id=EXCLUDED.product_id,
                     pattern=EXCLUDED.pattern, negate=EXCLUDED.negate, updated_at=now()
                   RETURNING id""",
                (name, f["modality_id"], f["product_id"], f["pattern"], f["negate"]),
            )
            sub_id[name] = c.fetchone()[0]

        n_att = 0
        for sname, subname in attach:
            c.execute(
                """INSERT INTO subscription_server (subscription_id, server_id)
                   VALUES (%s,%s) ON CONFLICT DO NOTHING""",
                (sub_id[subname], server_id[sname]),
            )
            n_att += c.rowcount
    g.commit()
    print(f"wrote {len(server_id)} servers, {len(sub_id)} subscriptions, {n_att} new attachment(s)")
    g.close()


if __name__ == "__main__":
    main()
